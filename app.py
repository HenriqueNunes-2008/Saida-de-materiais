from flask import Flask, render_template, request, redirect, url_for, flash  # type: ignore 
import csv
from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
from datetime import datetime
import os
from dotenv import load_dotenv

load_dotenv()
secret_key = os.getenv('SECRET_KEY')

app = Flask(__name__)

app.secret_key = secret_key

ARQUIVO_SAIDA = "saida_materiais.csv"
ARQUIVO_NOMES = "nomes.txt"
ARQUIVO_PROJETOS = "projetos.txt"

def obter_nome_arquivo_excel():
    agora = datetime.now()
    return f"saida_materiais_{agora.year}_{agora.month:02}.xlsx"

def carregar_lista(arquivo, padrao):
    if not os.path.exists(arquivo):
        with open(arquivo, "w", encoding="utf-8") as f:
            f.write("\n".join(padrao))
    with open(arquivo, "r", encoding="utf-8") as f:
        return [linha.strip() for linha in f if linha.strip()]

nomes = carregar_lista(ARQUIVO_NOMES, ["João", "Maria", "Carlos", "Fernanda"])
nomes.sort()
projetos = carregar_lista(ARQUIVO_PROJETOS, ["Projeto A", "Projeto B", "Projeto C"])
projetos.sort()

def registrar_em_excel(dados):
    nome_arquivo = obter_nome_arquivo_excel()

    if not os.path.exists(nome_arquivo):
        wb = Workbook()
        ws = wb.active
        ws.title = "Saídas"

        # Cabeçalhos
        headers = ["Data/Hora", "Nome", "Projeto", "Material", "Quantidade"]
        ws.append(headers)

        # Formatar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 20

    else:
        wb = load_workbook(nome_arquivo)
        ws = wb["Saídas"]

    nova_linha = ws.max_row + 1
    for col, valor in enumerate(dados, start=1):
        cell = ws.cell(row=nova_linha, column=col, value=valor)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    wb.save(nome_arquivo)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        nome = request.form.get("nome")
        projeto = request.form.get("projeto")
        material = request.form.get("material").strip()
        quantidade = request.form.get("quantidade").strip()

        if not (nome and projeto and material and quantidade):
            flash("Preencha todos os campos!", "error")
            return redirect(url_for("index"))

        try:
            quantidade_int = int(quantidade)
            if quantidade_int <= 0:
                raise ValueError
        except ValueError:
            flash("Quantidade deve ser um número inteiro positivo!", "error")
            return redirect(url_for("index"))

        data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        registro = [data, nome, projeto, material, quantidade]

        registrar_em_excel(registro)  # Grava no Excel

        flash("Saída registrada com sucesso!", "success")
        return redirect(url_for("index"))

    return render_template("index.html", nomes=nomes, projetos=projetos)

@app.route("/add_nome", methods=["POST"])
def add_nome():
    novo_nome = request.form.get("novo_nome", "").strip()
    if novo_nome:
        global nomes
        if novo_nome in nomes:
            flash("Esse nome já existe.", "error")
        else:
            with open(ARQUIVO_NOMES, "a", encoding="utf-8") as f:
                f.write(novo_nome + "\n")
            nomes.append(novo_nome)
            nomes.sort()
            flash(f"Nome '{novo_nome}' adicionado com sucesso!", "success")
    else:
        flash("Nome não pode ser vazio.", "error")
    return redirect(url_for("index"))

@app.route("/add_projeto", methods=["POST"])
def add_projeto():
    novo_projeto = request.form.get("novo_projeto", "").strip()
    if novo_projeto:
        global projetos
        if novo_projeto in projetos:
            flash("Esse projeto já existe.", "error")
        else:
            with open(ARQUIVO_PROJETOS, "a", encoding="utf-8") as f:
                f.write(novo_projeto + "\n")
            projetos.append(novo_projeto)
            projetos.sort()
            flash(f"Projeto '{novo_projeto}' adicionado com sucesso!", "success")
    else:
        flash("Projeto não pode ser vazio.", "error")
    return redirect(url_for("index"))

@app.route("/delete_nome/<nome>", methods=["POST"])
def delete_nome(nome):
    global nomes
    if nome in nomes:
        nomes.remove(nome)
        with open(ARQUIVO_NOMES, "w", encoding="utf-8") as f:
            for n in nomes:
                f.write(n + "\n")
        flash(f"Nome '{nome}' removido com sucesso!", "success")
    else:
        flash("Nome não encontrado.", "error")
    return redirect(url_for("view_lists"))

@app.route("/delete_projeto/<projeto>", methods=["POST"])
def delete_projeto(projeto):
    global projetos
    if projeto in projetos:
        projetos.remove(projeto)
        with open(ARQUIVO_PROJETOS, "w", encoding="utf-8") as f:
            for p in projetos:
                f.write(p + "\n")
        flash(f"Projeto '{projeto}' removido com sucesso!", "success")
    else:
        flash("Projeto não encontrado.", "error")
    return redirect(url_for("view_lists"))

@app.route("/view_lists")
def view_lists():
    return render_template("view_lists.html", nomes=nomes, projetos=projetos)

if __name__ == "__main__":
    app.run(debug=True)
